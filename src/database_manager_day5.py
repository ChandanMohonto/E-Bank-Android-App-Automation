"""
Day 5: Database Integration & Export System
Complete database management and export functionality for banking automation
"""

import sqlite3
import pandas as pd
import json
import csv
import logging
import time
from datetime import datetime, timedelta
from pathlib import Path
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DatabaseManager:
    def __init__(self, db_path="mobile_tests.db"):
        self.db_path = Path(db_path)
        self.logger = logging.getLogger(__name__)
        self.ensure_database_exists()
        
    def ensure_database_exists(self):
        """Create database and tables if they don't exist"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # UI Elements table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS ui_elements (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        scan_id TEXT NOT NULL,
                        app_name TEXT NOT NULL,
                        screen_name TEXT NOT NULL,
                        element_type TEXT NOT NULL,
                        element_id TEXT,
                        resource_id TEXT,
                        xpath TEXT,
                        accessibility_id TEXT,
                        text_content TEXT,
                        content_desc TEXT,
                        bounds TEXT,
                        clickable BOOLEAN,
                        enabled BOOLEAN,
                        displayed BOOLEAN,
                        password BOOLEAN,
                        class_name TEXT,
                        locator_strategies TEXT,  -- JSON string of locators
                        safety_level TEXT,
                        safety_reason TEXT,
                        automation_allowed BOOLEAN,
                        automation_notes TEXT,  -- JSON string of notes
                        detection_method TEXT,
                        screenshot_path TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(scan_id, resource_id, bounds)
                    )
                ''')
                
                # Test Cases table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS test_cases (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE,
                        description TEXT,
                        app_name TEXT NOT NULL,
                        platform TEXT NOT NULL DEFAULT 'Android',
                        status TEXT DEFAULT 'Draft',
                        priority TEXT DEFAULT 'Medium',
                        tags TEXT,  -- JSON array of tags
                        created_by TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                # Test Steps table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS test_steps (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        test_case_id INTEGER,
                        step_order INTEGER,
                        action_type TEXT NOT NULL,
                        locator_type TEXT,
                        locator_value TEXT,
                        input_data TEXT,
                        expected_result TEXT,
                        screenshot BOOLEAN DEFAULT FALSE,
                        wait_time INTEGER DEFAULT 1,
                        retry_count INTEGER DEFAULT 1,
                        safety_validation TEXT,  -- JSON of safety checks
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (test_case_id) REFERENCES test_cases (id) ON DELETE CASCADE
                    )
                ''')
                
                # Test Executions table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS test_executions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        test_case_id INTEGER,
                        execution_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        status TEXT NOT NULL,
                        total_steps INTEGER,
                        passed_steps INTEGER,
                        failed_steps INTEGER,
                        skipped_steps INTEGER,
                        execution_time_seconds REAL,
                        error_message TEXT,
                        environment_info TEXT,  -- JSON of device/app info
                        screenshot_dir TEXT,
                        execution_log TEXT,  -- Detailed log
                        FOREIGN KEY (test_case_id) REFERENCES test_cases (id)
                    )
                ''')
                
                # Scan Sessions table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS scan_sessions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        scan_id TEXT UNIQUE NOT NULL,
                        app_name TEXT NOT NULL,
                        screen_name TEXT NOT NULL,
                        scan_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        scan_duration REAL,
                        total_elements INTEGER,
                        device_info TEXT,  -- JSON of device information
                        scan_metadata TEXT,  -- JSON of scan results metadata
                        screenshot_path TEXT,
                        safety_warnings TEXT  -- JSON array of warnings
                    )
                ''')
                
                # Export History table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS export_history (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        export_type TEXT NOT NULL,
                        file_path TEXT NOT NULL,
                        export_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        record_count INTEGER,
                        filters_applied TEXT,  -- JSON of applied filters
                        file_size_bytes INTEGER,
                        exported_by TEXT
                    )
                ''')
                
                conn.commit()
                self.logger.info("Database schema initialized successfully")
                
        except Exception as e:
            self.logger.error(f"Failed to initialize database: {e}")
            raise
    
    def save_scan_results(self, scan_results):
        """
        Save complete scan results to database
        
        Args:
            scan_results: Dictionary containing scan results from element scanner
            
        Returns:
            str: Scan ID for the saved session
        """
        try:
            scan_id = f"scan_{int(time.time())}_{hashlib.md5(str(scan_results).encode()).hexdigest()[:8]}"
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Save scan session
                cursor.execute('''
                    INSERT INTO scan_sessions 
                    (scan_id, app_name, screen_name, scan_duration, total_elements, 
                     device_info, scan_metadata, screenshot_path, safety_warnings)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    scan_id,
                    scan_results.get('app_name', 'Unknown'),
                    scan_results.get('screen_name', 'Unknown'),
                    scan_results.get('scan_duration', 0),
                    len(scan_results.get('elements', [])),
                    json.dumps(scan_results.get('metadata', {})),
                    json.dumps(scan_results.get('statistics', {})),
                    scan_results.get('metadata', {}).get('screenshot_path', ''),
                    json.dumps(scan_results.get('warnings', []))
                ))
                
                # Save individual elements
                elements_saved = 0
                for element in scan_results.get('elements', []):
                    try:
                        safety_classification = element.get('safety_classification', {})
                        locators = element.get('locators', {})
                        automation_notes = element.get('automation_notes', [])
                        
                        cursor.execute('''
                            INSERT OR REPLACE INTO ui_elements 
                            (scan_id, app_name, screen_name, element_type, element_id, 
                             resource_id, xpath, accessibility_id, text_content, content_desc,
                             bounds, clickable, enabled, displayed, password, class_name,
                             locator_strategies, safety_level, safety_reason, automation_allowed,
                             automation_notes, detection_method, screenshot_path)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            scan_id,
                            scan_results.get('app_name', 'Unknown'),
                            scan_results.get('screen_name', 'Unknown'),
                            element.get('class_name', 'Unknown'),
                            element.get('resource_id', ''),
                            element.get('resource_id', ''),
                            locators.get('xpath_resource_id', ''),
                            element.get('content_desc', ''),
                            element.get('text', ''),
                            element.get('content_desc', ''),
                            element.get('bounds', ''),
                            element.get('clickable', False),
                            element.get('enabled', False),
                            element.get('displayed', False),
                            element.get('password', False),
                            element.get('class_name', ''),
                            json.dumps(locators),
                            safety_classification.get('level', 'UNKNOWN'),
                            safety_classification.get('reason', ''),
                            safety_classification.get('automation_allowed', True),
                            json.dumps(automation_notes),
                            element.get('detection_method', 'unknown'),
                            scan_results.get('metadata', {}).get('screenshot_path', '')
                        ))
                        elements_saved += 1
                        
                    except Exception as e:
                        self.logger.warning(f"Failed to save element: {e}")
                        continue
                
                conn.commit()
                self.logger.info(f"Saved scan session {scan_id} with {elements_saved} elements")
                return scan_id
                
        except Exception as e:
            self.logger.error(f"Failed to save scan results: {e}")
            return None
    
    def get_scan_sessions(self, limit=50, app_name=None):
        """Get recent scan sessions"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                query = '''
                    SELECT scan_id, app_name, screen_name, scan_timestamp, 
                           scan_duration, total_elements, safety_warnings
                    FROM scan_sessions
                '''
                params = []
                
                if app_name:
                    query += ' WHERE app_name = ?'
                    params.append(app_name)
                
                query += ' ORDER BY scan_timestamp DESC LIMIT ?'
                params.append(limit)
                
                cursor.execute(query, params)
                return cursor.fetchall()
                
        except Exception as e:
            self.logger.error(f"Failed to get scan sessions: {e}")
            return []
    
    def get_elements_by_scan(self, scan_id):
        """Get all elements from a specific scan"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT * FROM ui_elements WHERE scan_id = ?
                    ORDER BY created_at ASC
                ''', (scan_id,))
                
                columns = [description[0] for description in cursor.description]
                rows = cursor.fetchall()
                
                return [dict(zip(columns, row)) for row in rows]
                
        except Exception as e:
            self.logger.error(f"Failed to get elements for scan {scan_id}: {e}")
            return []
    
    def export_to_csv(self, output_path, filters=None):
        """
        Export UI elements to CSV with optional filtering
        
        Args:
            output_path: Path to save CSV file
            filters: Dictionary of filters to apply
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Build query with filters
                query = '''
                    SELECT 
                        scan_id,
                        app_name,
                        screen_name,
                        element_type,
                        resource_id,
                        text_content,
                        content_desc,
                        bounds,
                        clickable,
                        enabled,
                        safety_level,
                        safety_reason,
                        automation_allowed,
                        created_at
                    FROM ui_elements
                '''
                
                conditions = []
                params = []
                
                if filters:
                    if filters.get('app_name'):
                        conditions.append('app_name = ?')
                        params.append(filters['app_name'])
                    
                    if filters.get('safety_level'):
                        conditions.append('safety_level = ?')
                        params.append(filters['safety_level'])
                    
                    if filters.get('automation_allowed') is not None:
                        conditions.append('automation_allowed = ?')
                        params.append(filters['automation_allowed'])
                    
                    if filters.get('date_from'):
                        conditions.append('created_at >= ?')
                        params.append(filters['date_from'])
                
                if conditions:
                    query += ' WHERE ' + ' AND '.join(conditions)
                
                query += ' ORDER BY created_at DESC'
                
                # Execute query and save to CSV
                df = pd.read_sql_query(query, conn, params=params)
                df.to_csv(output_path, index=False)
                
                # Record export history
                self._record_export(output_path, 'CSV', len(df), filters)
                
                self.logger.info(f"Exported {len(df)} records to CSV: {output_path}")
                return len(df)
                
        except Exception as e:
            self.logger.error(f"CSV export failed: {e}")
            return 0
    
    def export_to_excel(self, output_path, filters=None, include_charts=True):
        """
        Export UI elements to Excel with formatting and optional charts
        
        Args:
            output_path: Path to save Excel file
            filters: Dictionary of filters to apply
            include_charts: Whether to include summary charts
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Get main data
                query = '''
                    SELECT 
                        scan_id,
                        app_name,
                        screen_name,
                        element_type,
                        resource_id,
                        text_content,
                        content_desc,
                        bounds,
                        clickable,
                        enabled,
                        displayed,
                        safety_level,
                        safety_reason,
                        automation_allowed,
                        automation_notes,
                        created_at
                    FROM ui_elements
                '''
                
                conditions = []
                params = []
                
                if filters:
                    if filters.get('app_name'):
                        conditions.append('app_name = ?')
                        params.append(filters['app_name'])
                    
                    if filters.get('safety_level'):
                        conditions.append('safety_level = ?')
                        params.append(filters['safety_level'])
                
                if conditions:
                    query += ' WHERE ' + ' AND '.join(conditions)
                
                query += ' ORDER BY created_at DESC'
                
                df = pd.read_sql_query(query, conn, params=params)
                
                # Create Excel workbook
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # Main data sheet
                    df.to_excel(writer, sheet_name='UI Elements', index=False)
                    
                    # Summary sheet
                    summary_data = self._create_summary_data(df)
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Safety analysis sheet
                    safety_analysis = df.groupby('safety_level').size().reset_index(name='count')
                    safety_analysis.to_excel(writer, sheet_name='Safety Analysis', index=False)
                    
                    # Format the main sheet
                    workbook = writer.book
                    worksheet = writer.sheets['UI Elements']
                    
                    # Header formatting
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for col_num, value in enumerate(df.columns.values):
                        cell = worksheet.cell(row=1, column=col_num + 1)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Color code safety levels
                    safety_colors = {
                        'HIGH_RISK': 'FFCCCC',  # Light red
                        'MEDIUM_RISK': 'FFFFCC',  # Light yellow
                        'LOW_RISK': 'CCFFCC',  # Light green
                        'SAFE': 'CCE5FF'  # Light blue
                    }
                    
                    safety_col = df.columns.get_loc('safety_level') + 1
                    for row_num in range(2, len(df) + 2):
                        safety_value = worksheet.cell(row=row_num, column=safety_col).value
                        if safety_value in safety_colors:
                            fill = PatternFill(start_color=safety_colors[safety_value], 
                                             end_color=safety_colors[safety_value], 
                                             fill_type="solid")
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row_num, column=col).fill = fill
                
                # Record export history
                self._record_export(output_path, 'Excel', len(df), filters)
                
                self.logger.info(f"Exported {len(df)} records to Excel: {output_path}")
                return len(df)
                
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            return 0
    
    def export_test_cases(self, output_path, format='excel'):
        """Export test cases and steps"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Get test cases with step counts
                test_cases_query = '''
                    SELECT 
                        tc.id,
                        tc.name,
                        tc.description,
                        tc.app_name,
                        tc.platform,
                        tc.status,
                        tc.priority,
                        tc.tags,
                        tc.created_at,
                        COUNT(ts.id) as step_count
                    FROM test_cases tc
                    LEFT JOIN test_steps ts ON tc.id = ts.test_case_id
                    GROUP BY tc.id
                    ORDER BY tc.created_at DESC
                '''
                
                test_cases_df = pd.read_sql_query(test_cases_query, conn)
                
                # Get test steps
                test_steps_query = '''
                    SELECT 
                        ts.*,
                        tc.name as test_case_name
                    FROM test_steps ts
                    JOIN test_cases tc ON ts.test_case_id = tc.id
                    ORDER BY tc.name, ts.step_order
                '''
                
                test_steps_df = pd.read_sql_query(test_steps_query, conn)
                
                if format.lower() == 'excel':
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        test_cases_df.to_excel(writer, sheet_name='Test Cases', index=False)
                        test_steps_df.to_excel(writer, sheet_name='Test Steps', index=False)
                else:
                    # CSV export (two files)
                    base_path = Path(output_path)
                    cases_path = base_path.with_suffix('.cases.csv')
                    steps_path = base_path.with_suffix('.steps.csv')
                    
                    test_cases_df.to_csv(cases_path, index=False)
                    test_steps_df.to_csv(steps_path, index=False)
                
                self._record_export(output_path, f'Test Cases ({format})', len(test_cases_df), {})
                
                self.logger.info(f"Exported {len(test_cases_df)} test cases to {format}: {output_path}")
                return len(test_cases_df)
                
        except Exception as e:
            self.logger.error(f"Test cases export failed: {e}")
            return 0
    
    def _create_summary_data(self, df):
        """Create summary statistics for export"""
        if df.empty:
            return []
        
        summary = [
            {'Metric': 'Total Elements', 'Value': len(df)},
            {'Metric': 'Unique Apps', 'Value': df['app_name'].nunique()},
            {'Metric': 'Unique Screens', 'Value': df['screen_name'].nunique()},
            {'Metric': 'Clickable Elements', 'Value': df['clickable'].sum()},
            {'Metric': 'Safe Elements', 'Value': (df['safety_level'] == 'SAFE').sum()},
            {'Metric': 'High Risk Elements', 'Value': (df['safety_level'] == 'HIGH_RISK').sum()},
            {'Metric': 'Automation Ready', 'Value': df['automation_allowed'].sum()},
        ]
        
        # Add date range
        if 'created_at' in df.columns:
            summary.extend([
                {'Metric': 'Date Range Start', 'Value': df['created_at'].min()},
                {'Metric': 'Date Range End', 'Value': df['created_at'].max()},
            ])
        
        return summary
    
    def _record_export(self, file_path, export_type, record_count, filters):
        """Record export operation in history"""
        try:
            file_size = Path(file_path).stat().st_size if Path(file_path).exists() else 0
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO export_history 
                    (export_type, file_path, record_count, filters_applied, file_size_bytes, exported_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    export_type,
                    str(file_path),
                    record_count,
                    json.dumps(filters or {}),
                    file_size,
                    'System'
                ))
                conn.commit()
                
        except Exception as e:
            self.logger.warning(f"Failed to record export history: {e}")
    
    def get_export_history(self, limit=20):
        """Get recent export history"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT export_type, file_path, export_timestamp, record_count, file_size_bytes
                    FROM export_history
                    ORDER BY export_timestamp DESC
                    LIMIT ?
                ''', (limit,))
                
                return cursor.fetchall()
                
        except Exception as e:
            self.logger.error(f"Failed to get export history: {e}")
            return []
    
    def cleanup_old_data(self, days_to_keep=30):
        """Clean up old scan data to keep database size manageable"""
        try:
            cutoff_date = datetime.now() - timedelta(days=days_to_keep)
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Delete old scan sessions and related elements
                cursor.execute('''
                    DELETE FROM ui_elements 
                    WHERE created_at < ?
                ''', (cutoff_date,))
                
                cursor.execute('''
                    DELETE FROM scan_sessions 
                    WHERE scan_timestamp < ?
                ''', (cutoff_date,))
                
                # Clean up old export history
                cursor.execute('''
                    DELETE FROM export_history 
                    WHERE export_timestamp < ?
                ''', (cutoff_date,))
                
                deleted_count = cursor.rowcount
                conn.commit()
                
                self.logger.info(f"Cleaned up {deleted_count} old records")
                return deleted_count
                
        except Exception as e:
            self.logger.error(f"Data cleanup failed: {e}")
            return 0
    
    def get_database_stats(self):
        """Get database statistics"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                stats = {}
                
                # Table counts
                tables = ['ui_elements', 'test_cases', 'test_steps', 'test_executions', 'scan_sessions']
                for table in tables:
                    cursor.execute(f'SELECT COUNT(*) FROM {table}')
                    stats[f'{table}_count'] = cursor.fetchone()[0]
                
                # Database size
                stats['database_size_mb'] = self.db_path.stat().st_size / (1024 * 1024)
                
                # Date ranges
                cursor.execute('SELECT MIN(created_at), MAX(created_at) FROM ui_elements')
                date_range = cursor.fetchone()
                stats['data_date_range'] = {
                    'start': date_range[0],
                    'end': date_range[1]
                }
                
                return stats
                
        except Exception as e:
            self.logger.error(f"Failed to get database stats: {e}")
            return {}

# Integration functions for main application
def save_scan_to_database(scan_results, db_path="mobile_tests.db"):
    """
    Convenience function to save scan results
    
    Args:
        scan_results: Scan results from element scanner
        db_path: Database file path
        
    Returns:
        str: Scan ID if successful, None if failed
    """
    try:
        db_manager = DatabaseManager(db_path)
        return db_manager.save_scan_results(scan_results)
    except Exception as e:
        logging.error(f"Failed to save scan to database: {e}")
        return None

def export_elements_to_file(output_path, format='csv', filters=None, db_path="mobile_tests.db"):
    """
    Convenience function to export elements
    
    Args:
        output_path: Output file path
        format: 'csv' or 'excel'
        filters: Export filters
        db_path: Database file path
        
    Returns:
        int: Number of records exported
    """
    try:
        db_manager = DatabaseManager(db_path)
        
        if format.lower() == 'csv':
            return db_manager.export_to_csv(output_path, filters)
        elif format.lower() == 'excel':
            return db_manager.export_to_excel(output_path, filters)
        else:
            raise ValueError(f"Unsupported format: {format}")
            
    except Exception as e:
        logging.error(f"Failed to export elements: {e}")
        return 0

if __name__ == "__main__":
    # Test the database manager
    logging.basicConfig(level=logging.INFO)
    
    db_manager = DatabaseManager("test_mobile.db")
    
    # Test data
    test_scan_results = {
        'app_name': 'Test Banking App',
        'screen_name': 'Login Screen',
        'scan_duration': 2.5,
        'elements': [
            {
                'class_name': 'android.widget.EditText',
                'resource_id': 'com.bank.app:id/username',
                'text': '',
                'content_desc': 'Username field',
                'bounds': '[100,200][400,250]',
                'clickable': True,
                'enabled': True,
                'displayed': True,
                'password': False,
                'locators': {
                    'resource_id': 'com.bank.app:id/username',
                    'xpath_resource_id': '//*[@resource-id="com.bank.app:id/username"]'
                },
                'safety_classification': {
                    'level': 'MEDIUM_RISK',
                    'reason': 'Login field requires caution',
                    'automation_allowed': True
                },
                'automation_notes': ['Use test credentials only'],
                'detection_method': 'resource_id'
            }
        ],
        'metadata': {
            'screenshot_path': '/screenshots/test.png'
        },
        'statistics': {
            'total_elements': 1
        },
        'warnings': ['Test warning']
    }
    
    # Test save
    scan_id = db_manager.save_scan_results(test_scan_results)
    print(f"Saved scan with ID: {scan_id}")
    
    # Test export
    exported = db_manager.export_to_csv("test_export.csv")
    print(f"Exported {exported} records to CSV")
    
    # Test stats
    stats = db_manager.get_database_stats()
    print(f"Database stats: {stats}")